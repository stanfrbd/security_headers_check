#!/usr/bin/env python3

# Stanislas M. 2023-03-17

"""
usage: security_headers_check_copy.py [-h] [-u URL] [-i INPUT_FILE] [-f {csv,xlsx}]

options:
  -h, --help            show this help message and exit
  -u URL, --url URL     http://example.com -- it will export to csv in the current directory
  -i INPUT_FILE, --input-file INPUT_FILE
                        Choose the path to input file containing URLs e.g. "test.txt" -- it will export to csv in the
                        current directory
  -f {csv,xlsx}, --format {csv,xlsx}
                        Choose the export format: csv (default) or xlsx
"""

import json
import argparse
import re
import requests
from bs4 import BeautifulSoup
from pathlib import Path
from datetime import datetime
import os
import sys
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo

# Your proxy here...
proxy = ""

# Current date
now = datetime.now()
today = now.strftime("%Y-%m-%d-%H_%M_%S")

# Required Headers
"""
| Required HTTP 1.1 (HTTPS):     |
| ------------------------------ |
| Content-Security-Policy        |
| HTTP Strict-Transport-Security |
| X-Content-Type-Options         |
| Cache-Control                  |

| Required HTTP 1.0 (HTTPS):     |
| ------------------------------ |
| Content-Security-Policy        |
| HTTP Strict-Transport-Security |
| X-Content-Type-Options         |
| Expires                        |
"""

REQUIRED_HEADERS_HTTP_10_OR_11 = [ "Cache-Control", "Content-Security-Policy", "Strict-Transport-Security", "X-Content-Type-Options", "Expires" ]

# Optional Headers
"""
| Header                      | HTTP Versions       |
| --------------------------- | ------------------- |
| Access-Control-Allow-Origin | HTTP/1.0 / HTTP/1.1 |
| Location                    | HTTP/1.0 / HTTP/1.1 |
| Set-Cookie                  | HTTP/1.0 / HTTP/1.1 |
| WWW-Authenticate            | HTTP/1.0 / HTTP/1.1 |
| X-Frame-Options             | HTTP/1.0 / HTTP/1.1 |
| X-XSS-Protection            | HTTP/1.0 / HTTP/1.1 |
| Permissions-Policy          | HTTP/1.0 / HTTP/1.1 |
| Referrer-Policy             | HTTP/1.0 / HTTP/1.1 |
"""

OPTIONAL_HEADERS = [ "Access-Control-Allow-Origin", "Location", "Set-Cookie", "WWW-Authenticate", "X-Frame-Options", "X-XSS-Protection", "Permissions-Policy", "Referrer-Policy" ]

# CHECKS

# Credit: https://www.adamsmith.haus/python/answers/how-to-remove-empty-lines-from-a-string-in-python
def remove_empty_lines(txt):
    lines = txt.split("\n")
    non_empty_lines = [line for line in lines if line.strip() != ""]
    string_without_empty_lines = ""
    for line in non_empty_lines:
      string_without_empty_lines += line.replace(";", " ") + " - "
    return string_without_empty_lines.replace(",", " ")

def export(data, file_format):
    if file_format == "csv":
        print("\nGenerated CSV: ./security-headers-" + today + "-export.csv\n")
        with open("security-headers-" + today + "-export.csv", "a") as f:
            f.write("site,score,httpforwarding,missing_required_headers,missing_optional_headers,warnings,security_headers_url\n")
            for row in data:
                f.write(",".join(row) + "\n")
    elif file_format == "xlsx":
        export_to_excel(data, "security-headers-" + today + "-export.xlsx")

def export_to_excel(data, filename):
    wb = Workbook()
    ws = wb.active
    ws.append(["site", "score", "httpforwarding", "missing_required_headers", "missing_optional_headers", "warnings", "security_headers_url"])

    # Define conditional formatting colors
    green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
    light_orange_fill = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")
    orange_fill = PatternFill(start_color="F4B183", end_color="F4B183", fill_type="solid")
    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    dark_gray_fill = PatternFill(start_color="A6A6A6", end_color="A6A6A6", fill_type="solid")

    # Add data to the worksheet
    for row in data:
        ws.append(row)

    # Apply conditional formatting
    for row in ws.iter_rows(min_row=2):
        # Score column (column B)
        if row[1].value.startswith("A"):
            row[1].fill = green_fill
        elif row[1].value.startswith("B"):
            row[1].fill = light_orange_fill
        elif row[1].value.startswith("C"):
            row[1].fill = orange_fill
        elif row[1].value.startswith("D") or row[1].value.startswith("E") or row[1].value.startswith("F"):
            row[1].fill = red_fill
        elif row[1].value == "Unknown":
            row[1].fill = dark_gray_fill

        # httpforwarding column (column C)
        if "amazon" in row[2].value:
            row[2].fill = orange_fill

        # missing_required_headers column (column D)
        if row[3].value == "No issue":
            row[3].fill = green_fill
        elif "|" in row[3].value:
            row[3].fill = red_fill

        # warnings column (column F)
        if "-" in row[5].value:
            row[5].fill = orange_fill

    # Add table with filters
    table = Table(displayName="ResultsTable", ref="A1:G{}".format(ws.max_row))
    table_style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    table.tableStyleInfo = table_style
    ws.add_table(table)

    # Save the workbook
    wb.save(filename)


def check_missing_required_headers(txt):
    missing_found = False
    found_missing_required_headers = ""
    for el in REQUIRED_HEADERS_HTTP_10_OR_11:
        if el in txt:
            missing_found = True
            found_missing_required_headers += el + " | "
    if missing_found == False:
        return "No issue"
    return found_missing_required_headers

def check_missing_optional_headers(txt):
    missing_found = False
    found_missing_optional_headers = ""
    for el in OPTIONAL_HEADERS:
        if el in txt:
            missing_found = True
            found_missing_optional_headers +=  el + " | "
    if missing_found == False:
        return "No issue"        
    return found_missing_optional_headers

def check_http_forwarding(url):
    try:
        proxy_servers = { 'http': proxy, 'https': proxy }
        if "http" not in url:
            url = "https://" + url
        response = requests.get(url, timeout=10, proxies=proxy_servers)
    except:
        return "Unreachable"
	
    try:
        if response.history:
            if len(response.history) == 1:
                # Redirect URL
                return response.url
            else:
                # Last forwarding URL
                return response.history[len(response.history) - 1].url
        else:
            return "No forwarding"
    except:
        return "Unknown"

# CORE FUNCTION

def scan(url):
    global csv
    
    data = []

    proxy_servers = { 'http': proxy, 'https': proxy }

    site = url
    score = ""
    httpforwarding = ""
    missing_req_headers = ""
    missing_opt_headers = ""
    warnings = ""
    securityheaders_url = "https://securityheaders.com/?q=" + url + "&followRedirects=on"
    base_request = requests.get(securityheaders_url, proxies=proxy_servers)

    if base_request.status_code == 200:
        base_text = base_request.text

        soup = BeautifulSoup(base_text, "html.parser")
        print("\nTitle: ", soup.title.string) 

        # https://securityheaders.com
        try:
            # Score
            score_div = soup.find_all("div", class_="score")
            score = re.search("[A-Z]", str(score_div[0])).group()

            # Dividing content in sections
            report_sections = soup.find_all("div", class_="reportSection")

            # Missing Headers or Warnings
            extracted_data3 = remove_empty_lines(report_sections[3].get_text())
            if "Missing Headers" in extracted_data3:
                missing_req_headers = check_missing_required_headers(extracted_data3)
                missing_opt_headers = check_missing_optional_headers(extracted_data3)
            elif "Warnings" in extracted_data3:
                warnings = extracted_data3
                missing_req_headers = "No issue"
                missing_opt_headers = "No issue"

            # Warnings if exist
            extracted_data4 = remove_empty_lines(report_sections[4].get_text())
            if "Warnings" in extracted_data4:
                warnings = extracted_data4
            else:
                warnings = ""

        except Exception: 
            score = "Unknown"
            missing_req_headers = "Unknown"
            missing_opt_headers = "Unknown"
            warnings = "Unknown"

        # HTTP Forwarding
        httpforwarding = check_http_forwarding(url)

        print("Score: ", score)
        print("URL:   ", securityheaders_url)
        print("HTTP Forwarding: ", httpforwarding)

    else:
        raise Exception("HTTP error: " + str(base_request.status_code))
    
    return [site, score, httpforwarding, missing_req_headers, missing_opt_headers, warnings, securityheaders_url]

# URL -u / --url

def action_url(txt, export_format):
    print("# ", txt)
    data = [scan(txt)]
    export(data, export_format)

# INPUT_FILE -i / --input-file

def action_file(txt, export_format):
    all_data = []

    if Path(txt).is_file():
        input_file = open(txt, "r")
        data = input_file.readlines()
        input_file.close()
        if data != []:
            for i in range (0, len(data)):
                output_data = scan(data[i].rstrip('\n'))
                all_data.append(output_data)    
    else:
        print("\"" + txt + "\" is not a valid file, aborting.")
    
    export(all_data, export_format)

# MAIN

def main(): 
    parser = argparse.ArgumentParser()
    parser.add_argument('-u','--url', help='http://example.com -- it will export to csv in the current directory')
    parser.add_argument('-i','--input-file', help='Choose the path to input file containing URLs e.g. "test.txt" -- it will export to csv in the current directory')
    parser.add_argument('-f', '--format', choices=['csv', 'xlsx'], default='csv', help='Choose the export format: csv (default) or xlsx')
    
    global args
    args = parser.parse_args()

    if args.url:
        action_url(args.url, args.format)

    if args.input_file:
        action_file(args.input_file, args.format)     

if __name__ == "__main__":
    try: 
        main() 
    except Exception as err: 
        print("General error: ", err) 
        exit(1)
